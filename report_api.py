"""
QRIE Report Generation API
Generates well-formatted reports from PNB scanner data in multiple formats.
No database required — reads directly from public/data/PNB/ JSON files.
"""

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel, Field
from typing import List, Optional
from pathlib import Path
from datetime import datetime, timezone
import json, csv, io, os, uuid, tempfile, shutil

# ── PDF / XLSX imports (deferred so server still boots if missing) ────────────
try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm, cm
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, HRFlowable, Image, KeepTogether
    )
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.graphics.shapes import Drawing, String, Rect
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.legends import Legend
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─────────────────────────────────────────────────────────────────────────────
# App setup
# ─────────────────────────────────────────────────────────────────────────────
app = FastAPI(title="QRIE Report API", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

DATA_DIR = Path(__file__).parent / "public" / "data" / "PNB"
REPORTS_DIR = Path(__file__).parent / "generated_reports"
SCHEDULES_FILE = Path(__file__).parent / "report_schedules.json"
REPORTS_DIR.mkdir(exist_ok=True)

# ── PNB Brand colours ────────────────────────────────────────────────────────
PNB_CRIMSON = (139, 0, 0)
PNB_AMBER   = (217, 119, 6)
PNB_GOLD    = (245, 158, 11)
PNB_CREAM   = (255, 248, 231)

# ─────────────────────────────────────────────────────────────────────────────
# Pydantic models
# ─────────────────────────────────────────────────────────────────────────────
class ReportRequest(BaseModel):
    report_types: List[str] = Field(..., description="List of report type keys")
    format: str = Field("pdf", description="pdf | json | csv | xlsx | cyclonedx")
    include_charts: bool = True
    password_protect: bool = False
    sections: Optional[List[str]] = None  # optional section filter

class EmailRequest(ReportRequest):
    recipients: List[str] = Field(..., description="Email addresses")

class ScheduleRequest(ReportRequest):
    frequency: str = Field("Weekly", description="Daily | Weekly | Bi-Weekly | Monthly | Quarterly")
    schedule_date: str = Field(..., description="ISO date string")
    schedule_time: str = Field("09:00 AM (IST)")
    assets_scope: str = Field("All Assets")
    delivery_email: Optional[str] = None
    delivery_save_path: Optional[str] = None
    delivery_link: bool = False
    enabled: bool = True

class LinkRequest(ReportRequest):
    pass

# ─────────────────────────────────────────────────────────────────────────────
# Data loaders
# ─────────────────────────────────────────────────────────────────────────────
def _load_json(filename: str):
    fp = DATA_DIR / filename
    if not fp.exists():
        return None
    with open(fp, "r", encoding="utf-8") as f:
        return json.load(f)

def _dedupe_by_domain(records: list) -> list:
    """Keep one canonical record per unique Asset (domain)."""
    m = {}
    for r in records:
        key = r.get("Asset", "")
        if key not in m:
            m[key] = r
        else:
            existing = m[key]
            r_ok = r.get("Scan Status") == "ok"
            ex_ok = existing.get("Scan Status") == "ok"
            if r_ok and r.get("Port") == 443:
                m[key] = r
            elif r_ok and not ex_ok:
                m[key] = r
    return list(m.values())

# ─────────────────────────────────────────────────────────────────────────────
# Report data aggregators — one per report type
# ─────────────────────────────────────────────────────────────────────────────
def _fmt_type(key):
    return (key or "Unknown").replace("_", " ").title()

def _agg_executive():
    """Executive Summary: KPIs, risk distribution, cert expiry, infra."""
    raw = _load_json("enriched_cbom.json")
    if not raw:
        return {"title": "Executive Summary", "error": "No data available"}
    summary = raw.get("_PQC_Enrichment_Summary", {})
    records = _dedupe_by_domain(raw.get("records", []))
    total = len(records)

    risk_counts = {"Critical": 0, "Legacy": 0, "Standard": 0, "Elite": 0}
    cert_expiry = {"Expired": 0, "0-30 Days": 0, "30-90 Days": 0, ">90 Days": 0}
    now = datetime.now(timezone.utc)

    for r in records:
        hei = r.get("HEI_Score", 50)
        if hei < 20:   risk_counts["Elite"] += 1
        elif hei < 50: risk_counts["Standard"] += 1
        elif hei < 80: risk_counts["Legacy"] += 1
        else:          risk_counts["Critical"] += 1

        cv = r.get("Certificate Validity (Not Before/After)", {})
        na = cv.get("Not After")
        if na:
            try:
                if na.endswith("Z"):
                    na = na[:-1] + "+00:00"
                exp = datetime.fromisoformat(na)
                if exp.tzinfo is None:
                    exp = exp.replace(tzinfo=timezone.utc)
                days = (exp - now).days
                if days < 0:       cert_expiry["Expired"] += 1
                elif days <= 30:   cert_expiry["0-30 Days"] += 1
                elif days <= 90:   cert_expiry["30-90 Days"] += 1
                else:              cert_expiry[">90 Days"] += 1
            except:
                pass

    at_dist = summary.get("asset_type_distribution", {})
    infra = summary.get("infrastructure_summary", {})

    return {
        "title": "Executive Summary Report",
        "generated_at": datetime.now().isoformat(),
        "organization": "Punjab National Bank",
        "kpis": {
            "total_assets": total,
            "total_records": summary.get("total_assets", total),
            "web_applications": (at_dist.get("web_application", 0) + at_dist.get("web_server", 0)),
            "apis_detected": at_dist.get("api", 0),
            "cdn_proxy": at_dist.get("cdn_proxy", 0),
            "avg_hei_score": summary.get("HEI", {}).get("avg", 0),
            "min_hei_score": summary.get("HEI", {}).get("min", 0),
            "max_hei_score": summary.get("HEI", {}).get("max", 0),
        },
        "risk_distribution": risk_counts,
        "certificate_expiry": cert_expiry,
        "asset_type_distribution": at_dist,
        "infrastructure_summary": infra,
        "cipher_strength_distribution": summary.get("cipher_strength_distribution", {}),
        "qrmm_distribution": summary.get("qrmm_distribution", {}),
        "cert_distribution": summary.get("cert_distribution", {}),
        "top_priority_assets": summary.get("top10_by_priority", [])[:5],
    }

def _agg_discovery():
    """Asset Discovery: subdomains, IPs, domains."""
    subs_raw = _load_json("subdomains.json")
    cbom_raw = _load_json("enriched_cbom.json")
    subs = subs_raw if isinstance(subs_raw, list) else (subs_raw or {}).get("subdomains", [])
    records = _dedupe_by_domain((cbom_raw or {}).get("records", []))

    # Unique IPs
    ips = set()
    for r in records:
        ip = r.get("IP Address", "")
        if ip and ip != "—":
            ips.add(ip)

    # SSL certificates
    ssl_certs = []
    seen = set()
    for r in records:
        asset = r.get("Asset", "")
        if r.get("Issuer CA") and asset not in seen:
            seen.add(asset)
            cv = r.get("Certificate Validity (Not Before/After)", {})
            ssl_d = r.get("SSL Details", {})
            ssl_certs.append({
                "asset": asset,
                "issuer": (r.get("Issuer CA", "")).replace("CN=", ""),
                "valid_from": cv.get("Not Before", "—"),
                "valid_to": cv.get("Not After", "—"),
                "days_left": ssl_d.get("days_until_expiry"),
                "cipher_strength": ssl_d.get("cipher_strength", "Unknown"),
                "is_ev": ssl_d.get("is_ev", False),
                "is_wildcard": ssl_d.get("is_wildcard", False),
                "protocol": ssl_d.get("protocol_version") or r.get("TLS Version", "—"),
            })

    return {
        "title": "Asset Discovery Report",
        "generated_at": datetime.now().isoformat(),
        "organization": "Punjab National Bank",
        "total_subdomains": len(subs),
        "subdomains": subs[:200] if (subs and isinstance(subs[0], str)) else [],
        "unique_ips": sorted(list(ips)),
        "total_unique_ips": len(ips),
        "ssl_certificates": ssl_certs[:100],
        "total_ssl_certs": len(ssl_certs),
    }

def _agg_inventory():
    """Asset Inventory: full CBOM records with TLS/crypto details."""
    raw = _load_json("enriched_cbom.json")
    if not raw:
        return {"title": "Asset Inventory Report", "error": "No data"}
    records = _dedupe_by_domain(raw.get("records", []))
    items = []
    for r in records:
        cv = r.get("Certificate Validity (Not Before/After)", {})
        items.append({
            "asset": r.get("Asset", "Unknown"),
            "ip": r.get("IP Address", "—"),
            "port": r.get("Port"),
            "asset_type": _fmt_type(r.get("Asset Type", "unknown")),
            "tls_version": r.get("TLS Version", "—"),
            "cipher_suite": r.get("Cipher Suite", "—"),
            "key_exchange": r.get("Key Exchange Algorithm", "—"),
            "key_size": r.get("Key Size (Bits)", "—"),
            "pfs_status": r.get("PFS Status", "Unknown"),
            "issuer_ca": (r.get("Issuer CA") or "—").replace("CN=", ""),
            "valid_to": cv.get("Not After", "—"),
            "scan_status": r.get("Scan Status", "—"),
            "hei_score": r.get("HEI_Score", "—"),
            "risk_category": r.get("Risk_Category", "—"),
        })
    return {
        "title": "Asset Inventory Report",
        "generated_at": datetime.now().isoformat(),
        "organization": "Punjab National Bank",
        "total_assets": len(items),
        "assets": items,
    }

def _agg_cbom():
    """CBOM: Cryptographic Bill of Materials."""
    raw = _load_json("enriched_cbom.json")
    if not raw:
        return {"title": "CBOM Report", "error": "No data"}
    summary = raw.get("_PQC_Enrichment_Summary", {})
    records = _dedupe_by_domain(raw.get("records", []))

    cipher_counts = {}
    ca_counts = {}
    tls_counts = {}
    items = []

    for r in records:
        cipher = r.get("Cipher Suite") or "Unknown"
        ca = (r.get("Issuer CA") or "Other").replace("CN=", "").split(",")[0].strip()
        tls = r.get("TLS Version") or "Unknown"
        cipher_counts[cipher] = cipher_counts.get(cipher, 0) + 1
        ca_counts[ca] = ca_counts.get(ca, 0) + 1
        tls_counts[tls] = tls_counts.get(tls, 0) + 1

        ssl_d = r.get("SSL Details", {})
        items.append({
            "asset": r.get("Asset", "Unknown"),
            "cipher_suite": cipher,
            "key_size": f"{r.get('Key Size (Bits)', '—')}-bit",
            "tls_version": tls,
            "issuer_ca": ca[:30],
            "cipher_strength": ssl_d.get("cipher_strength", "Unknown"),
            "is_ev": ssl_d.get("is_ev", False),
            "is_wildcard": ssl_d.get("is_wildcard", False),
            "weak": (
                (r.get("Key Size (Bits)") or 2048) < 2048
                or "1.0" in tls or "1.1" in tls
                or "DES" in cipher
            ),
        })

    return {
        "title": "CBOM Report",
        "generated_at": datetime.now().isoformat(),
        "organization": "Punjab National Bank",
        "total_assets": len(items),
        "cipher_distribution": dict(sorted(cipher_counts.items(), key=lambda x: -x[1])[:10]),
        "ca_distribution": dict(sorted(ca_counts.items(), key=lambda x: -x[1])[:10]),
        "tls_distribution": dict(sorted(tls_counts.items(), key=lambda x: -x[1])[:5]),
        "cipher_strength_distribution": summary.get("cipher_strength_distribution", {}),
        "assets": items,
    }

def _agg_pqc():
    """PQC Posture: quantum readiness grades."""
    raw = _load_json("enriched_cbom.json")
    if not raw:
        return {"title": "PQC Posture Report", "error": "No data"}
    records = _dedupe_by_domain(raw.get("records", []))
    summary = raw.get("_PQC_Enrichment_Summary", {})

    elite = std = legacy = critical = 0
    items = []
    for r in records:
        hei = r.get("HEI_Score", 50)
        pqc_label = r.get("NIST PQC Readiness Label", "")
        is_pqc = "PQC" in pqc_label or hei < 20
        if hei < 20:   elite += 1
        elif hei < 50: std += 1
        elif hei < 80: legacy += 1
        else:          critical += 1

        qrmm = r.get("QRMM_Level", {})
        items.append({
            "asset": r.get("Asset", "Unknown"),
            "ip": r.get("IP Address", "—"),
            "hei_score": hei,
            "risk_category": r.get("Risk_Category", "—"),
            "pqc_ready": is_pqc,
            "pqc_label": pqc_label,
            "qrmm_level": qrmm.get("level", "—"),
            "qrmm_label": qrmm.get("label", "—"),
            "mds_score": r.get("MDS_Score", "—"),
            "certification_status": r.get("Certification_Status", "—"),
        })

    total = len(records) or 1
    return {
        "title": "PQC Posture Report",
        "generated_at": datetime.now().isoformat(),
        "organization": "Punjab National Bank",
        "total_assets": len(items),
        "grade_summary": {
            "Elite-PQC Ready": elite,
            "Standard": std,
            "Legacy": legacy,
            "Critical": critical,
        },
        "percentages": {
            "pqc_ready_pct": round((elite / total) * 100, 1),
            "standard_pct": round((std / total) * 100, 1),
            "legacy_pct": round((legacy / total) * 100, 1),
            "critical_pct": round((critical / total) * 100, 1),
        },
        "qrmm_distribution": summary.get("qrmm_distribution", {}),
        "cert_distribution": summary.get("cert_distribution", {}),
        "hei_summary": summary.get("HEI", {}),
        "mds_summary": summary.get("MDS", {}),
        "assets": items,
    }

def _agg_cyber_rating():
    """Cyber Rating: per-asset scores and enterprise tier."""
    raw = _load_json("enriched_cbom.json")
    if not raw:
        return {"title": "Cyber Rating Report", "error": "No data"}
    records = _dedupe_by_domain(raw.get("records", []))
    total_score = 0
    items = []
    for r in records:
        hei = r.get("HEI_Score", 50)
        score = max(0, 1000 - (hei * 10))
        total_score += score
        tier = "Elite" if score >= 701 else "Standard" if score >= 400 else "Legacy" if score >= 200 else "Critical"
        items.append({
            "asset": r.get("Asset", "Unknown"),
            "score": round(score),
            "tier": tier,
            "hei_score": hei,
            "asset_type": _fmt_type(r.get("Asset Type", "unknown")),
        })

    items.sort(key=lambda x: -x["score"])
    enterprise_score = round(total_score / len(records)) if records else 0
    enterprise_tier = (
        "Elite-PQC" if enterprise_score >= 701
        else "Standard" if enterprise_score >= 400
        else "Legacy" if enterprise_score >= 200
        else "Critical"
    )

    # Tier distribution
    tier_dist = {"Elite": 0, "Standard": 0, "Legacy": 0, "Critical": 0}
    for it in items:
        tier_dist[it["tier"]] = tier_dist.get(it["tier"], 0) + 1

    return {
        "title": "Cyber Rating Report",
        "generated_at": datetime.now().isoformat(),
        "organization": "Punjab National Bank",
        "enterprise_score": enterprise_score,
        "enterprise_tier": enterprise_tier,
        "total_assets": len(items),
        "tier_distribution": tier_dist,
        "assets": items[:100],
    }

# ─────────────────────────────────────────────────────────────────────────────
# Aggregator dispatcher
# ─────────────────────────────────────────────────────────────────────────────
REPORT_TYPE_MAP = {
    "Executive Summary Report":   _agg_executive,
    "Executive Reporting":        _agg_executive,
    "executive":                  _agg_executive,
    "Asset Discovery Report":     _agg_discovery,
    "Assets Discovery":           _agg_discovery,
    "discovery":                  _agg_discovery,
    "Asset Inventory Report":     _agg_inventory,
    "Assets Inventory":           _agg_inventory,
    "inventory":                  _agg_inventory,
    "CBOM Report":                _agg_cbom,
    "CBOM":                       _agg_cbom,
    "cbom":                       _agg_cbom,
    "PQC Posture Report":         _agg_pqc,
    "Posture of PQC":             _agg_pqc,
    "pqc":                        _agg_pqc,
    "Cyber Rating Report":        _agg_cyber_rating,
    "Cyber Rating (Tiers 1 - 4)": _agg_cyber_rating,
    "cyber_rating":               _agg_cyber_rating,
}

def _collect_report_data(report_types: List[str]) -> List[dict]:
    """Run aggregators for each requested report type."""
    sections = []
    for rt in report_types:
        agg = REPORT_TYPE_MAP.get(rt)
        if agg:
            sections.append(agg())
        else:
            sections.append({"title": rt, "error": f"Unknown report type: {rt}"})
    return sections


# ─────────────────────────────────────────────────────────────────────────────
# Format generators
# ─────────────────────────────────────────────────────────────────────────────

def _generate_json(sections: list, filepath: Path):
    """Write a JSON report."""
    report = {
        "report_metadata": {
            "generated_at": datetime.now().isoformat(),
            "organization": "Punjab National Bank",
            "generator": "QRIE Report Engine v1.0",
            "total_sections": len(sections),
        },
        "sections": sections,
    }
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, default=str)

def _generate_csv(sections: list, filepath: Path):
    """Write a CSV report (flattened tables)."""
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["QRIE Report — Punjab National Bank"])
        writer.writerow([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        writer.writerow([])

        for section in sections:
            title = section.get("title", "Section")
            writer.writerow([f"═══ {title} ═══"])
            writer.writerow([])

            # Write summary KPIs first
            for key, val in section.items():
                if key in ("title", "assets", "subdomains", "ssl_certificates",
                           "unique_ips", "top_priority_assets", "generated_at",
                           "organization", "error"):
                    continue
                if isinstance(val, dict):
                    writer.writerow([key.replace("_", " ").title()])
                    for k2, v2 in val.items():
                        writer.writerow(["", k2, v2])
                    writer.writerow([])
                elif not isinstance(val, list):
                    writer.writerow([key.replace("_", " ").title(), val])

            # Write asset table if present
            assets = section.get("assets", [])
            if assets and isinstance(assets[0], dict):
                writer.writerow([])
                headers = list(assets[0].keys())
                writer.writerow(headers)
                for item in assets[:500]:
                    writer.writerow([item.get(h, "") for h in headers])

            # Write subdomains if present
            subs = section.get("subdomains", [])
            if subs:
                writer.writerow([])
                writer.writerow(["Subdomains"])
                for s in subs:
                    writer.writerow([s if isinstance(s, str) else s.get("fqdn", "")])

            writer.writerow([])
            writer.writerow([])

def _generate_xlsx(sections: list, filepath: Path):
    """Write an XLSX report with PNB-branded styling."""
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="openpyxl not installed — XLSX generation unavailable")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    subheader_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
    subheader_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    border = Border(
        left=Side(style="thin", color="D4D4D4"),
        right=Side(style="thin", color="D4D4D4"),
        top=Side(style="thin", color="D4D4D4"),
        bottom=Side(style="thin", color="D4D4D4"),
    )

    for idx, section in enumerate(sections):
        title = section.get("title", f"Section {idx+1}")
        # Sheet name max 31 chars
        sheet_name = title[:31].replace("/", "-")
        ws = wb.create_sheet(title=sheet_name)

        row = 1
        # Title row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=f"QRIE — {title}")
        cell.font = Font(name="Calibri", bold=True, size=14, color="8B0000")
        row += 1

        ws.cell(row=row, column=1, value=f"Organization: Punjab National Bank")
        ws.cell(row=row, column=4, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        row += 2

        # Summary KPIs
        for key, val in section.items():
            if key in ("title", "assets", "subdomains", "ssl_certificates",
                        "unique_ips", "top_priority_assets", "generated_at",
                        "organization", "error"):
                continue
            if isinstance(val, dict):
                c = ws.cell(row=row, column=1, value=key.replace("_", " ").title())
                c.font = subheader_font
                c.fill = subheader_fill
                ws.cell(row=row, column=2).fill = subheader_fill
                ws.cell(row=row, column=3).fill = subheader_fill
                row += 1
                for k2, v2 in val.items():
                    ws.cell(row=row, column=2, value=str(k2)).font = data_font
                    ws.cell(row=row, column=3, value=str(v2)).font = data_font
                    row += 1
                row += 1
            elif not isinstance(val, list):
                ws.cell(row=row, column=1, value=key.replace("_", " ").title()).font = Font(bold=True, size=10)
                ws.cell(row=row, column=2, value=str(val)).font = data_font
                row += 1

        row += 1

        # Asset data table
        assets = section.get("assets", [])
        if assets and isinstance(assets[0], dict):
            headers = list(assets[0].keys())
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=ci, value=h.replace("_", " ").title())
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center")
                c.border = border
            row += 1

            for item in assets[:500]:
                for ci, h in enumerate(headers, 1):
                    c = ws.cell(row=row, column=ci, value=str(item.get(h, "")))
                    c.font = data_font
                    c.border = border
                row += 1

        # Auto-width
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    wb.save(filepath)

def _generate_pdf(sections: list, filepath: Path):
    """Generate a professional PNB-branded PDF report."""
    if not HAS_REPORTLAB:
        raise HTTPException(status_code=500, detail="reportlab not installed — PDF generation unavailable")

    doc = SimpleDocTemplate(
        str(filepath),
        pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=25*mm, bottomMargin=20*mm,
    )

    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        "PNBTitle", parent=styles["Title"],
        fontSize=22, textColor=rl_colors.HexColor("#8B0000"),
        spaceAfter=6*mm, fontName="Helvetica-Bold",
    )
    heading_style = ParagraphStyle(
        "PNBHeading", parent=styles["Heading1"],
        fontSize=16, textColor=rl_colors.HexColor("#8B0000"),
        spaceBefore=8*mm, spaceAfter=4*mm, fontName="Helvetica-Bold",
    )
    subheading_style = ParagraphStyle(
        "PNBSubheading", parent=styles["Heading2"],
        fontSize=12, textColor=rl_colors.HexColor("#D97706"),
        spaceBefore=4*mm, spaceAfter=2*mm, fontName="Helvetica-Bold",
    )
    body_style = ParagraphStyle(
        "PNBBody", parent=styles["Normal"],
        fontSize=9, textColor=rl_colors.HexColor("#333333"),
        spaceAfter=2*mm, fontName="Helvetica",
    )
    kpi_style = ParagraphStyle(
        "PNBKPI", parent=styles["Normal"],
        fontSize=10, textColor=rl_colors.HexColor("#1a1a1a"),
        spaceAfter=1*mm, fontName="Helvetica",
    )

    elements = []

    # ── Cover section ──
    elements.append(Spacer(1, 40*mm))
    
    # Large Title Block
    title_data = [
        [Paragraph("<b>QRIE</b>", ParagraphStyle(
            "C_Title_Main", fontSize=36, textColor=rl_colors.white, fontName="Helvetica-Bold", alignment=TA_LEFT
        ))],
        [Paragraph("Quantum Risk Intelligence Engine", ParagraphStyle(
            "C_Title_Sub1", fontSize=14, textColor=rl_colors.HexColor("#F59E0B"), fontName="Helvetica", alignment=TA_LEFT, spaceBefore=2*mm
        ))],
        [Paragraph("PUNJAB NATIONAL BANK", ParagraphStyle(
            "C_Title_Sub2", fontSize=16, textColor=rl_colors.white, fontName="Helvetica-Bold", alignment=TA_LEFT, spaceBefore=6*mm
        ))]
    ]
    title_tbl = Table(title_data, colWidths=[170*mm])
    title_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), rl_colors.HexColor("#8B0000")),
        ('TOPPADDING', (0,0), (-1,-1), 15*mm),
        ('BOTTOMPADDING', (0,0), (-1,-1), 15*mm),
        ('LEFTPADDING', (0,0), (-1,-1), 10*mm),
    ]))
    elements.append(title_tbl)
    elements.append(Spacer(1, 25*mm))
    
    # Subtitle / Report contents
    section_titles = " • ".join(s.get("title", "Report") for s in sections)
    elements.append(Paragraph("DOCUMENT CONTENTS", ParagraphStyle(
        "C_Doc_Cont", fontSize=10, textColor=rl_colors.HexColor("#D97706"), fontName="Helvetica-Bold", alignment=TA_LEFT, spaceAfter=2*mm
    )))
    elements.append(Paragraph(section_titles, ParagraphStyle(
        "C_Doc_Cont_Val", fontSize=14, textColor=rl_colors.HexColor("#1A1A1A"), fontName="Helvetica", alignment=TA_LEFT, spaceAfter=20*mm, leading=18
    )))
    
    # Metadata Block
    meta_data = [
        [
            Paragraph("<b>Generated On:</b>", ParagraphStyle("M_Lbl", fontSize=10, textColor=rl_colors.gray, fontName="Helvetica")),
            Paragraph(datetime.now().strftime('%B %d, %Y at %H:%M IST'), ParagraphStyle("M_Val", fontSize=10, textColor=rl_colors.black, fontName="Helvetica-Bold"))
        ],
        [
            Paragraph("<b>Classification:</b>", ParagraphStyle("M_Lbl", fontSize=10, textColor=rl_colors.gray, fontName="Helvetica")),
            Paragraph("<font color='#CC0000'>STRICTLY CONFIDENTIAL</font>", ParagraphStyle("M_Val", fontSize=10, textColor=rl_colors.black, fontName="Helvetica-Bold"))
        ],
        [
            Paragraph("<b>Generated By:</b>", ParagraphStyle("M_Lbl", fontSize=10, textColor=rl_colors.gray, fontName="Helvetica")),
            Paragraph("QRIE Automated Reporting Service", ParagraphStyle("M_Val", fontSize=10, textColor=rl_colors.black, fontName="Helvetica"))
        ]
    ]
    meta_tbl = Table(meta_data, colWidths=[35*mm, 135*mm])
    meta_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(meta_tbl)
    
    elements.append(PageBreak())

    # ── Each section ──
    for si, section in enumerate(sections):
        if si > 0:
            elements.append(PageBreak())

        title = section.get("title", "Report Section")
        elements.append(Paragraph(title, heading_style))
        elements.append(HRFlowable(
            width="100%", thickness=1,
            color=rl_colors.HexColor("#D97706"), spaceAfter=4*mm,
        ))

        if section.get("error"):
            elements.append(Paragraph(f"Error: {section['error']}", body_style))
            continue

        def _build_pie(data_dict):
            valid_data = {str(k): float(v) for k,v in data_dict.items() if isinstance(v, (int, float)) and v > 0}
            if not valid_data: return None
            
            sorted_data = sorted(valid_data.items(), key=lambda x: -x[1])
            
            # Group into Top 8 + "Other" to prevent legend spillage
            MAX_ITEMS = 7
            if len(sorted_data) > (MAX_ITEMS + 1):
                top_items = sorted_data[:MAX_ITEMS]
                other_sum = sum(v for k, v in sorted_data[MAX_ITEMS:])
                top_items.append(("Other", other_sum))
                sorted_data = top_items
                
            d = Drawing(160*mm, 60*mm)
            pc = Pie()
            pc.x = 10*mm; pc.y = 5*mm; pc.width = 50*mm; pc.height = 50*mm
            
            pc.data = [v for k,v in sorted_data]
            # No labels on the pie slices to avoid overlap, legend handles it
            pc.labels = []
            
            colors_list = [rl_colors.HexColor(c) for c in ["#8B0000","#D97706","#F59E0B","#1D4ED8","#047857","#6B7280","#4338CA","#BE123C"]]
            for i in range(len(pc.data)):
                pc.slices[i].fillColor = colors_list[i % len(colors_list)]
                pc.slices[i].popout = 3 if i == 0 else 0
            
            d.add(pc)
            
            leg = Legend()
            leg.fontName = "Helvetica"
            leg.fontSize = 8
            leg.alignment = 'right'
            leg.x = 75*mm; leg.y = 50*mm
            leg.yGap = 3
            leg.dxTextSpace = 8
            
            def _trunc(s):
                # We have more room since it's one chart per row
                return s[:50] + '...' if len(s) > 50 else s
                
            leg.colorNamePairs = [(colors_list[i % len(colors_list)], f"{_trunc(sorted_data[i][0])} ({sorted_data[i][1]:n})") for i in range(len(pc.data))]
            d.add(leg)
            return d

        # KPIs / summary dict fields
        kpis = []
        charts = []
        tables = []
        
        for key, val in section.items():
            if key in ("title", "assets", "subdomains", "ssl_certificates", "unique_ips", "top_priority_assets", "generated_at", "organization", "error"):
                continue

            label = key.replace("_", " ").title()

            if isinstance(val, dict):
                # Check if it looks like a distribution
                is_num_dict = all(isinstance(v, (int,float)) for v in val.values())
                if is_num_dict and len(val) > 0 and ("distribution" in key.lower() or "summary" in key.lower() or "expiry" in key.lower() or "cert_expiry" in key.lower()):
                    chart = _build_pie(val)
                    if chart:
                        charts.append((label, chart))
                        continue
                
                # Render as mini table if not a chart
                tdata = [["Metric", "Value"]]
                for k2, v2 in val.items():
                    tdata.append([str(k2), str(v2)])
                if len(tdata) > 1:
                    t = Table(tdata, colWidths=[90*mm, 70*mm], hAlign='LEFT')
                    t.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#D97706")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#D4D4D4")),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.HexColor("#FFF8E7"), rl_colors.white]),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]))
                    tables.append((label, t))
            elif not isinstance(val, list):
                kpis.append((label, str(val)))
                
        # Render KPIs as a grid
        if kpis:
            kpi_data = []
            row = []
            for lbl, v in kpis:
                # We'll make sub-tables for each KPI to look like a card
                card_data = [[Paragraph(f"<font size='8' color='#666666'>{lbl}</font>")], [Paragraph(f"<font size='14' color='#8B0000'><b>{v}</b></font>")]]
                card = Table(card_data, colWidths=[50*mm])
                card.setStyle(TableStyle([
                    ("BACKGROUND", (0,0), (-1,-1), rl_colors.HexColor("#FAFAFA")),
                    ("BOX", (0,0), (-1,-1), 0.5, rl_colors.HexColor("#E5E7EB")),
                    ("TOPPADDING", (0,0), (-1,-1), 4),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                ]))
                row.append(card)
                if len(row) == 3:
                    kpi_data.append(row)
                    row = []
            if row:
                while len(row) < 3: row.append("")
                kpi_data.append(row)
            
            elements.append(Paragraph("Key Performance Indicators", subheading_style))
            kpi_grid = Table(kpi_data, colWidths=[55*mm, 55*mm, 55*mm], hAlign='LEFT')
            kpi_grid.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP")]))
            elements.append(kpi_grid)
            elements.append(Spacer(1, 4*mm))
            
        # Render Charts - One per line
        if charts:
            for lbl, chart in charts:
                c_tbl = Table([
                    [Paragraph(lbl, subheading_style)],
                    [chart]
                ], colWidths=[160*mm], hAlign='LEFT')
                elements.append(KeepTogether([c_tbl, Spacer(1, 6*mm)]))
            
        # Render Tables
        for lbl, tbl in tables:
            elements.append(KeepTogether([
                Paragraph(lbl, subheading_style),
                Spacer(1, 1*mm),
                tbl,
                Spacer(1, 4*mm)
            ]))

        # Asset table
        assets = section.get("assets", [])
        if assets and isinstance(assets[0], dict):
            elements.append(Spacer(1, 4*mm))
            elements.append(Paragraph("Detailed Asset Data", subheading_style))

            headers = list(assets[0].keys())
            # Limit columns for readability
            display_headers = headers[:8]
            tdata = [[h.replace("_", " ").title()[:15] for h in display_headers]]

            for item in assets[:60]:
                row = []
                for h in display_headers:
                    v = str(item.get(h, ""))[:25]
                    row.append(v)
                tdata.append(row)

            col_w = 170*mm / len(display_headers)
            t = Table(tdata, colWidths=[col_w] * len(display_headers))
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#8B0000")),
                ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 6.5),
                ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.HexColor("#D4D4D4")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [rl_colors.HexColor("#FFF8E7"), rl_colors.white]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ]))
            elements.append(t)

    # Build
    def _footer(canvas, doc):
        canvas.saveState()
        # Footer line
        canvas.setStrokeColor(rl_colors.HexColor("#D97706"))
        canvas.setLineWidth(0.5)
        canvas.line(20*mm, 15*mm, A4[0]-20*mm, 15*mm)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(rl_colors.HexColor("#666666"))
        canvas.drawString(20*mm, 11*mm, "QRIE — Quantum Risk Intelligence Engine | Punjab National Bank | CONFIDENTIAL")
        canvas.drawRightString(A4[0]-20*mm, 11*mm, f"Page {doc.page}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)

def _generate_cyclonedx(sections: list, filepath: Path):
    """Generate a CycloneDX BOM (JSON format)."""
    # Collect all assets across all sections that have crypto data
    components = []
    raw = _load_json("enriched_cbom.json")
    records = _dedupe_by_domain((raw or {}).get("records", []))

    for r in records:
        comp = {
            "type": "cryptographic-asset",
            "name": r.get("Asset", "Unknown"),
            "version": r.get("TLS Version", "unknown"),
            "bom-ref": r.get("Asset ID", str(uuid.uuid4())),
            "properties": [
                {"name": "ip_address", "value": r.get("IP Address", "")},
                {"name": "port", "value": str(r.get("Port", ""))},
                {"name": "cipher_suite", "value": r.get("Cipher Suite") or ""},
                {"name": "key_exchange", "value": r.get("Key Exchange Algorithm") or ""},
                {"name": "key_size_bits", "value": str(r.get("Key Size (Bits)") or "")},
                {"name": "pfs_status", "value": r.get("PFS Status", "")},
                {"name": "pqc_readiness", "value": r.get("NIST PQC Readiness Label", "")},
                {"name": "hei_score", "value": str(r.get("HEI_Score", ""))},
                {"name": "risk_category", "value": r.get("Risk_Category", "")},
                {"name": "issuer_ca", "value": r.get("Issuer CA") or ""},
                {"name": "signature_algorithm", "value": r.get("Signature Algorithm") or ""},
            ],
        }
        components.append(comp)

    bom = {
        "bomFormat": "CycloneDX",
        "specVersion": "1.5",
        "serialNumber": f"urn:uuid:{uuid.uuid4()}",
        "version": 1,
        "metadata": {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "tools": [{"vendor": "QRIE", "name": "Report Engine", "version": "1.0.0"}],
            "component": {
                "type": "application",
                "name": "Punjab National Bank — Cryptographic Inventory",
                "version": "1.0",
            },
        },
        "components": components,
    }
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(bom, f, indent=2, default=str)


FORMAT_GENERATORS = {
    "pdf":       (_generate_pdf,       ".pdf",  "application/pdf"),
    "json":      (_generate_json,      ".json", "application/json"),
    "csv":       (_generate_csv,       ".csv",  "text/csv"),
    "xlsx":      (_generate_xlsx,      ".xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    "cyclonedx": (_generate_cyclonedx, ".json", "application/json"),
}


def _build_report(req: ReportRequest) -> Path:
    """Aggregate data, generate file, return path."""
    fmt = req.format.lower()
    if fmt not in FORMAT_GENERATORS:
        raise HTTPException(400, f"Unsupported format: {fmt}. Use: {list(FORMAT_GENERATORS.keys())}")

    gen_fn, ext, _ = FORMAT_GENERATORS[fmt]
    sections = _collect_report_data(req.report_types)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    type_slug = "_".join(rt.split()[0].lower() for rt in req.report_types[:3])
    filename = f"QRIE_{type_slug}_{ts}{ext}"
    filepath = REPORTS_DIR / filename
    gen_fn(sections, filepath)
    return filepath


# ─────────────────────────────────────────────────────────────────────────────
# Endpoints
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/reports/health")
async def health():
    return {"status": "ok", "message": "QRIE Report API running",
            "has_reportlab": HAS_REPORTLAB, "has_openpyxl": HAS_OPENPYXL}

@app.get("/api/reports/formats")
async def list_formats():
    return {
        "formats": [
            {"key": "pdf",       "label": "PDF",        "ext": ".pdf",  "available": HAS_REPORTLAB},
            {"key": "json",      "label": "JSON",       "ext": ".json", "available": True},
            {"key": "csv",       "label": "CSV",        "ext": ".csv",  "available": True},
            {"key": "xlsx",      "label": "XLSX",       "ext": ".xlsx", "available": HAS_OPENPYXL},
            {"key": "cyclonedx", "label": "CycloneDX",  "ext": ".json", "available": True},
        ],
        "report_types": [
            "Executive Summary Report",
            "Asset Discovery Report",
            "Asset Inventory Report",
            "CBOM Report",
            "PQC Posture Report",
            "Cyber Rating Report",
        ],
    }

@app.post("/api/reports/generate")
async def generate_report(req: ReportRequest):
    """Generate a report and return it as a file download."""
    filepath = _build_report(req)
    _, _, media_type = FORMAT_GENERATORS[req.format.lower()]
    return FileResponse(
        path=str(filepath),
        filename=filepath.name,
        media_type=media_type,
    )

@app.post("/api/reports/email")
async def email_report(req: EmailRequest):
    """Generate a report and simulate email delivery."""
    filepath = _build_report(req)
    return {
        "success": True,
        "message": f"Report generated and sent to: {', '.join(req.recipients)}",
        "filename": filepath.name,
        "recipients": req.recipients,
        "format": req.format,
        "report_types": req.report_types,
        "note": "Email delivery simulated — configure SMTP for production use.",
    }

@app.post("/api/reports/link")
async def create_download_link(req: LinkRequest):
    """Generate a report and return a download link."""
    filepath = _build_report(req)
    link_id = str(uuid.uuid4())[:8]

    # Store link mapping
    links_file = REPORTS_DIR / "_links.json"
    links = {}
    if links_file.exists():
        with open(links_file) as f:
            links = json.load(f)
    links[link_id] = {
        "filename": filepath.name,
        "created_at": datetime.now().isoformat(),
        "format": req.format,
    }
    with open(links_file, "w") as f:
        json.dump(links, f, indent=2)

    return {
        "success": True,
        "link_id": link_id,
        "download_url": f"/api/reports/download/{link_id}",
        "filename": filepath.name,
    }

@app.get("/api/reports/download/{link_id}")
async def download_by_link(link_id: str):
    """Download a previously generated report by link ID."""
    links_file = REPORTS_DIR / "_links.json"
    if not links_file.exists():
        raise HTTPException(404, "No download links found")
    with open(links_file) as f:
        links = json.load(f)
    if link_id not in links:
        raise HTTPException(404, "Download link not found or expired")

    filename = links[link_id]["filename"]
    filepath = REPORTS_DIR / filename
    if not filepath.exists():
        raise HTTPException(404, "Report file not found")

    fmt = links[link_id].get("format", "json")
    _, _, media_type = FORMAT_GENERATORS.get(fmt, (None, None, "application/octet-stream"))
    return FileResponse(path=str(filepath), filename=filename, media_type=media_type)

@app.post("/api/reports/schedule")
async def save_schedule(req: ScheduleRequest):
    """Save a scheduled report configuration."""
    schedules = []
    if SCHEDULES_FILE.exists():
        with open(SCHEDULES_FILE) as f:
            schedules = json.load(f)

    schedule_entry = {
        "id": str(uuid.uuid4())[:8],
        "created_at": datetime.now().isoformat(),
        "report_types": req.report_types,
        "format": req.format,
        "frequency": req.frequency,
        "schedule_date": req.schedule_date,
        "schedule_time": req.schedule_time,
        "assets_scope": req.assets_scope,
        "sections": req.sections,
        "include_charts": req.include_charts,
        "delivery": {
            "email": req.delivery_email,
            "save_path": req.delivery_save_path,
            "link": req.delivery_link,
        },
        "enabled": req.enabled,
    }
    schedules.append(schedule_entry)

    with open(SCHEDULES_FILE, "w") as f:
        json.dump(schedules, f, indent=2)

    return {
        "success": True,
        "message": f"Schedule created: {req.frequency} {', '.join(req.report_types)}",
        "schedule": schedule_entry,
    }

@app.get("/api/reports/schedules")
async def list_schedules():
    """List all saved schedules."""
    if not SCHEDULES_FILE.exists():
        return {"schedules": []}
    with open(SCHEDULES_FILE) as f:
        schedules = json.load(f)
    return {"schedules": schedules}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8002)
